import openpyxl

# Открываем файл Excel
wb = openpyxl.load_workbook('C:\AAA\1.xlsx')

# Выбираем активный лист
ws = wb.active

# Определяем максимальное количество строк
max_row = ws.max_row

# Определяем количество групп по 200
group_count = max_row // 200 + (max_row % 200 > 0)

# Проходим по всем группам и записываем их в столбец B
for i in range(group_count):
    start_row = i * 200 + 1
    end_row = (i + 1) * 200
    group = '; '.join([str(ws.cell(row=row, column=1).value) for row in range(start_row, min(end_row, max_row) + 1)])
    ws.cell(row=i + 1, column=2).value = group

# Сохраняем изменения в файл
wb.save('C:\AAA\1.xlsx')
